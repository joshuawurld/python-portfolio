"""One-off script to generate XLSX test fixtures (run once, commit outputs).

Produces:
  fixtures/file_merged.xlsx  — an XLSX with merged cells to test warning detection
  fixtures/file_multisheet.xlsx — an XLSX with two sheets
"""
from pathlib import Path

from openpyxl import Workbook

FIX = Path(__file__).parent / "fixtures"


def make_merged():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Amount", "Currency"])
    ws.append([1, "Alpha", 100, "GBP"])
    ws.append([2, "Bravo", 200, "GBP"])
    ws.merge_cells("C2:D2")
    wb.save(FIX / "file_merged.xlsx")
    print("Created file_merged.xlsx")


def make_multisheet():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Value"])
    ws1.append([1, "Alpha", 100])
    ws1.append([2, "Bravo", 200])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name", "Value"])
    ws2.append([1, "Alpha", 999])
    ws2.append([2, "Bravo", 200])
    wb.save(FIX / "file_multisheet.xlsx")
    print("Created file_multisheet.xlsx")


if __name__ == "__main__":
    make_merged()
    make_multisheet()
